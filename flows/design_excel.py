"""Phase 1 of `SchemaDesignFlow`: deterministic Excel parse + FK graph + clustering.

Pure Python, no LLM calls. Handles the Excel layout from the project brief::

    Entity Name | Field Name | Field Full Name | Field Definition |
    Data Type   | Primary Key | Foreign Key

Sheets are scanned column-by-column with case-insensitive header matching so
mildly different exports still parse. Merged "Entity Name" cells are forward
filled so each row knows which entity it belongs to.

The FK graph and cluster partition are computed deterministically so 100+
table workbooks stay stable across runs. The optional ``python-louvain`` /
``networkx`` dependencies are imported lazily; missing libraries fall back to
trivial degree-bucket / "one big cluster" partitions so unit tests stay simple.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Iterable, Optional

from models.design_models import (
    ClusterSpec,
    ParsedField,
    ParsedSchema,
    ParsedTable,
)

logger = logging.getLogger(__name__)

# Canonical header names, lowercased & punctuation-stripped for matching.
_HEADER_ALIASES = {
    "entity_name": {"entity name", "entity", "table", "table name"},
    "field_name": {"field name", "field", "column", "column name"},
    "field_full_name": {"field full name", "full name", "label", "display name"},
    "field_definition": {"field definition", "definition", "description", "desc"},
    "data_type": {"data type", "type", "pg type", "datatype", "sql type"},
    "primary_key": {"primary key", "pk", "is pk", "is primary key"},
    "foreign_key": {"foreign key", "fk", "references", "ref", "reference"},
}

# A value in the PK column counts as "is PK" if it lower-cases to any of:
_PK_TRUE_VALUES = {"y", "yes", "true", "1", "t", "pk"}


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _match_header(value: object) -> Optional[str]:
    text = _norm_header(value)
    if not text:
        return None
    for canonical, aliases in _HEADER_ALIASES.items():
        if text == canonical.replace("_", " "):
            return canonical
        if text in aliases:
            return canonical
    return None


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _is_pk(value: object) -> bool:
    text = _stringify(value).lower()
    if not text:
        return False
    return text in _PK_TRUE_VALUES


def _clean_fk(value: object) -> Optional[str]:
    """Normalize a Foreign Key cell into ``Table.Field`` form.

    The hand-written dictionaries we see in the wild mix several conventions:
    ``Party.PartyId``, ``party.party_id``, ``party(party_id)`` (SQL DDL
    style), ``account_id_mapping -> accounts(account)``, and noise like
    ``code and become primary key )``. We normalize what we can and only
    return a best-effort string; ``_resolve_fk_target`` still has the final
    say on whether a target table actually exists.
    """
    text = _stringify(value)
    if not text:
        return None
    # Normalize arrow / dash separators.
    text = text.replace("→", ".").replace("->", ".").replace(" - ", ".")
    # Strip leading "ref:" / "fk:" prefixes some dictionaries use.
    text = re.sub(r"^(?:ref|references|fk)\s*[:=]\s*", "", text, flags=re.I)
    # Convert SQL-DDL style ``table(col)`` -> ``table.col`` (we only care
    # about the first paren group; anything after is treated as commentary).
    m = re.match(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*\(\s*([^),]+?)\s*\)", text)
    if m:
        text = f"{m.group(1)}.{m.group(2)}"
    # If, after normalization, the value still looks like a free-text comment
    # (no dot, lots of spaces, stray closing paren), keep just the first
    # whitespace-delimited token so the resolver can do a single table lookup.
    if " " in text and "." not in text:
        text = text.split(maxsplit=1)[0].rstrip(")")
    return text.strip()


def _parse_sheet(sheet, sheet_name: str) -> list[ParsedTable]:
    """Parse a single openpyxl worksheet into a list of `ParsedTable`."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Find the first non-empty row -- that is the header row.
    header_idx = -1
    for i, row in enumerate(rows):
        if any(cell not in (None, "") for cell in row):
            header_idx = i
            break
    if header_idx < 0:
        return []

    header_row = rows[header_idx]
    column_map: dict[int, str] = {}
    for col_idx, cell in enumerate(header_row):
        matched = _match_header(cell)
        if matched:
            column_map[col_idx] = matched

    if "entity_name" not in column_map.values() or "field_name" not in column_map.values():
        logger.warning(
            "design_excel: sheet '%s' missing required headers (entity/field); skipping",
            sheet_name,
        )
        return []

    tables: dict[str, ParsedTable] = {}
    last_entity: str = ""

    for row in rows[header_idx + 1 :]:
        if not any(cell not in (None, "") for cell in row):
            continue
        record: dict[str, object] = {}
        for col_idx, key in column_map.items():
            if col_idx < len(row):
                record[key] = row[col_idx]

        entity_raw = _stringify(record.get("entity_name"))
        if entity_raw:
            last_entity = entity_raw
        entity = last_entity
        if not entity:
            continue

        field_name = _stringify(record.get("field_name"))
        if not field_name:
            continue

        field = ParsedField(
            name=field_name,
            full_name=_stringify(record.get("field_full_name")),
            definition=_stringify(record.get("field_definition")),
            data_type=_stringify(record.get("data_type")),
            primary_key=_is_pk(record.get("primary_key")),
            foreign_key=_clean_fk(record.get("foreign_key")),
        )

        table = tables.get(entity)
        if table is None:
            table = ParsedTable(entity_name=entity, source_sheet=sheet_name)
            tables[entity] = table
        table.fields.append(field)

    return list(tables.values())


def parse_excel(content: bytes, filename: str = "schema.xlsx") -> ParsedSchema:
    """Parse the uploaded workbook into a :class:`ParsedSchema`.

    ``content`` is the raw bytes from the multipart upload. ``filename`` is
    used only for diagnostics.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency required at runtime
        raise RuntimeError(
            "openpyxl is required to parse uploaded workbooks"
        ) from exc

    wb = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    all_tables: dict[str, ParsedTable] = {}
    sheet_count = 0

    for sheet in wb.worksheets:
        sheet_count += 1
        parsed = _parse_sheet(sheet, sheet.title)
        for table in parsed:
            existing = all_tables.get(table.entity_name)
            if existing is None:
                all_tables[table.entity_name] = table
            else:
                # Merge: keep the existing source_sheet, append fields by name.
                seen = {f.name for f in existing.fields}
                for f in table.fields:
                    if f.name not in seen:
                        existing.fields.append(f)
                        seen.add(f.name)

    tables_list = list(all_tables.values())
    fk_count = sum(1 for t in tables_list for f in t.fields if f.foreign_key)

    logger.info(
        "design_excel: parsed %s -> %d tables, %d sheets, %d FK fields",
        filename,
        len(tables_list),
        sheet_count,
        fk_count,
    )

    return ParsedSchema(
        tables=tables_list,
        sheet_count=sheet_count,
        fk_count=fk_count,
    )


def _resolve_fk_target(raw: str, table_names: set[str]) -> Optional[tuple[str, str]]:
    """Return (target_table, target_field) if the FK string can be matched."""
    if not raw:
        return None
    text = raw.strip()
    # Support "Table.Field", "Table:Field", "Table Field".
    for sep in (".", ":", " "):
        if sep in text:
            a, b = text.split(sep, 1)
            a = a.strip()
            b = b.strip()
            if a in table_names:
                return (a, b)
            # Case-insensitive lookup.
            lowered = {n.lower(): n for n in table_names}
            if a.lower() in lowered:
                return (lowered[a.lower()], b)
    # No separator -> assume it's the target table referencing its PK.
    if text in table_names:
        return (text, "")
    lowered = {n.lower(): n for n in table_names}
    if text.lower() in lowered:
        return (lowered[text.lower()], "")
    return None


def build_fk_edges(schema: ParsedSchema) -> list[dict]:
    """Return a list of FK edges as plain dicts (used by ERDLayout.edges)."""
    table_names = {t.entity_name for t in schema.tables}
    edges: list[dict] = []
    for table in schema.tables:
        for field in table.fields:
            if not field.foreign_key:
                continue
            target = _resolve_fk_target(field.foreign_key, table_names)
            if target is None:
                continue
            edges.append(
                {
                    "from_table": table.entity_name,
                    "to_table": target[0],
                    "from_field": field.name,
                    "to_field": target[1],
                }
            )
    return edges


def _louvain_clusters(table_names: list[str], edges: list[dict]) -> dict[str, int]:
    """Try to compute a Louvain partition. Falls back to trivial partitions
    when dependencies are missing or the graph is too small.
    """
    try:
        import networkx as nx
    except ImportError:
        logger.info(
            "design_excel: networkx not installed; using trivial single-cluster partition"
        )
        return {name: 0 for name in table_names}

    g = nx.Graph()
    g.add_nodes_from(table_names)
    for edge in edges:
        a = edge.get("from_table")
        b = edge.get("to_table")
        if not a or not b or a == b:
            continue
        if g.has_edge(a, b):
            g[a][b]["weight"] = g[a][b].get("weight", 1) + 1
        else:
            g.add_edge(a, b, weight=1)

    if g.number_of_edges() == 0:
        # No FK relationships -> one cluster per connected component (just isolated nodes).
        return {name: 0 for name in table_names}

    try:
        import community as community_louvain  # python-louvain

        partition = community_louvain.best_partition(g, random_state=42)
        # Ensure tables with no edges still appear in partition map.
        for name in table_names:
            partition.setdefault(name, max(partition.values(), default=-1) + 1)
        return partition
    except ImportError:
        logger.info(
            "design_excel: python-louvain not installed; using connected-components partition"
        )
        cluster_of: dict[str, int] = {}
        for idx, component in enumerate(nx.connected_components(g)):
            for node in component:
                cluster_of[node] = idx
        next_idx = max(cluster_of.values(), default=-1) + 1
        for name in table_names:
            if name not in cluster_of:
                cluster_of[name] = next_idx
                next_idx += 1
        return cluster_of


def _looks_like_lookup(table: ParsedTable) -> bool:
    """A reference/lookup table is small and has no outgoing FKs.

    Heuristic only -- the analyst agent can override the categorization
    later. Used to fold many singleton reference tables into a single
    'Reference Data' cluster so the analyst sees ~5 meaningful clusters
    instead of 20+ trivial ones.
    """
    if len(table.fields) > 5:
        return False
    outgoing = sum(1 for f in table.fields if f.foreign_key)
    return outgoing == 0


def _consolidate_singletons(
    partition: dict[str, int],
    schema: ParsedSchema,
) -> dict[str, int]:
    """Fold singleton lookup-shaped tables into one ``Reference Data`` bucket.

    Louvain on a sparse / partially-disconnected FK graph routinely emits a
    swarm of size-1 clusters that are clearly reference data (TypeList,
    CountryCode, etc.). Merging them into one bucket lets the LLM treat
    them uniformly without producing dozens of tiny ``Cluster N`` blobs.
    Non-lookup singletons keep their own cluster -- they may be genuinely
    orphan business entities the analyst should flag.
    """
    counts: dict[int, int] = {}
    for cid in partition.values():
        counts[cid] = counts.get(cid, 0) + 1

    by_name = {t.entity_name: t for t in schema.tables}
    reference_bucket = max(partition.values(), default=-1) + 1
    moved = 0
    for name, cid in list(partition.items()):
        if counts.get(cid, 0) != 1:
            continue
        table = by_name.get(name)
        if table is None:
            continue
        if _looks_like_lookup(table):
            partition[name] = reference_bucket
            moved += 1

    if moved == 0:
        return partition

    # Re-number cluster ids so they are contiguous 0..N-1 and stable.
    ordered: dict[int, int] = {}
    next_id = 0
    for cid in sorted(set(partition.values())):
        ordered[cid] = next_id
        next_id += 1
    return {name: ordered[cid] for name, cid in partition.items()}


def build_clusters(schema: ParsedSchema, edges: Iterable[dict] | None = None) -> list[ClusterSpec]:
    """Deterministic clustering of tables. Names are placeholders -- the
    `DomainAnalystAgent` renames them later.
    """
    table_names = [t.entity_name for t in schema.tables]
    if not table_names:
        return []

    edge_list = list(edges) if edges is not None else build_fk_edges(schema)
    partition = _louvain_clusters(table_names, edge_list)
    partition = _consolidate_singletons(partition, schema)

    buckets: dict[int, list[str]] = {}
    for name, cid in partition.items():
        buckets.setdefault(cid, []).append(name)

    by_name = {t.entity_name: t for t in schema.tables}
    specs: list[ClusterSpec] = []
    for idx, (_raw_id, members) in enumerate(sorted(buckets.items())):
        members_sorted = sorted(members)
        is_reference = all(
            _looks_like_lookup(by_name[m]) for m in members_sorted if m in by_name
        ) and len(members_sorted) > 1
        specs.append(
            ClusterSpec(
                cluster_id=f"c{idx}",
                name="Reference Data" if is_reference else f"Cluster {idx + 1}",
                table_names=members_sorted,
                rationale=(
                    "Lookup/reference tables consolidated by deterministic heuristic"
                    if is_reference
                    else "Deterministic FK-graph community (Louvain)."
                ),
            )
        )
    return specs


def parse_and_cluster(content: bytes, filename: str = "schema.xlsx") -> tuple[ParsedSchema, list[ClusterSpec], list[dict]]:
    """Convenience: parse the workbook, build FK edges, build clusters.

    Returns ``(parsed_schema, clusters, fk_edges)``.
    """
    schema = parse_excel(content, filename=filename)
    edges = build_fk_edges(schema)
    clusters = build_clusters(schema, edges=edges)
    return schema, clusters, edges


def compute_layout(
    table_names: list[str],
    edges: list[dict],
    clusters: list[ClusterSpec],
) -> list[dict]:
    """Return a list of ``{table_name, x, y, z, cluster_id}`` dicts.

    Layout strategy: cluster tables on a ring, then spread tables inside each
    cluster on a smaller ring. Pure-Python so we do not need a JS layout pass on
    the backend. The frontend can re-layout client-side if desired.
    """
    import math

    cluster_of: dict[str, str] = {}
    for cluster in clusters:
        for name in cluster.table_names:
            cluster_of[name] = cluster.cluster_id

    cluster_ids = [c.cluster_id for c in clusters] or ["c0"]
    cluster_centers: dict[str, tuple[float, float, float]] = {}
    big_r = max(8.0, len(cluster_ids) * 2.5)
    for i, cid in enumerate(cluster_ids):
        theta = (2 * math.pi * i) / max(1, len(cluster_ids))
        cluster_centers[cid] = (math.cos(theta) * big_r, 0.0, math.sin(theta) * big_r)

    members: dict[str, list[str]] = {cid: [] for cid in cluster_ids}
    for name in table_names:
        cid = cluster_of.get(name, cluster_ids[0])
        members.setdefault(cid, []).append(name)

    layout: list[dict] = []
    for cid, names in members.items():
        cx, cy, cz = cluster_centers.get(cid, (0.0, 0.0, 0.0))
        n = len(names)
        small_r = max(1.5, n * 0.35)
        for j, name in enumerate(sorted(names)):
            phi = (2 * math.pi * j) / max(1, n)
            x = cx + math.cos(phi) * small_r
            z = cz + math.sin(phi) * small_r
            y = cy + (j % 3 - 1) * 0.6
            layout.append(
                {
                    "table_name": name,
                    "x": round(x, 3),
                    "y": round(y, 3),
                    "z": round(z, 3),
                    "cluster_id": cid,
                }
            )
    return layout
